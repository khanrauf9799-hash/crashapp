import os
import json
import eventlet
eventlet.monkey_patch()

from flask import Flask, render_template, jsonify, request, send_file, Response
from flask_socketio import SocketIO, emit
from datetime import datetime, timedelta
import threading
import time
import io

from sqlalchemy import text

from models import (db, Round, Pattern, PatternStats, StreakHistory, Alert, 
                   PatternSubStats, ColorConfig, NotificationSettings, get_color_from_value, 
                   invalidate_color_config_cache)
from pattern_engine import PatternEngine, DEFAULT_PATTERNS
from bc_game_listener import BCGameListener
import requests as http_requests

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'crash-pattern-engine-secret')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///crash.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')

pattern_engine = PatternEngine()
bc_listener = None
pattern_stats_cache = {}
recent_colors = []
recent_rounds = []
MAX_COLORS_BUFFER = 500
alert_cooldown_cache = {}
DEFAULT_TRIGGER_THRESHOLD = 5
DEFAULT_MISS_THRESHOLD = 2
DEFAULT_COOLDOWN_ROUNDS = 1


def _is_valid_count_expr(value):
    if isinstance(value, int):
        return value > 0
    if isinstance(value, str):
        v = value.strip()
        if v.isdigit():
            return int(v) > 0
        if v.startswith('$X'):
            rest = v[2:]
            if rest == '':
                return True
            if rest.startswith(('+', '-')):
                return rest[1:].strip().isdigit()
    return False


def _validate_rule_json(rule):
    if not isinstance(rule, dict):
        return False, 'rule_json must be an object'

    if rule.get('type') not in ['static', 'dynamic']:
        return False, 'type must be static or dynamic'

    if not isinstance(rule.get('name', ''), str) or not rule.get('name', '').strip():
        return False, 'name is required'

    allowed_colors = {'red', 'green', 'yellow', 'green_yellow'}
    allowed_ops = {'exact', 'minimum', 'maximum'}

    def _validate_steps(path):
        steps = rule.get(path, {}).get('steps', [])
        if not isinstance(steps, list):
            return False, f'{path}.steps must be a list'
        for s in steps:
            if not isinstance(s, dict):
                return False, f'{path}.steps contains non-object step'
            colors = s.get('colors', [])
            if not isinstance(colors, list) or not colors:
                return False, f'{path}.steps colors required'
            if any(c not in allowed_colors for c in colors):
                return False, f'{path}.steps contains invalid colors'
            if s.get('operator', 'minimum') not in allowed_ops:
                return False, f'{path}.steps contains invalid operator'
            if not _is_valid_count_expr(s.get('count', 1)):
                return False, f'{path}.steps contains invalid count'
        return True, None

    ok, err = _validate_steps('trigger')
    if not ok:
        return False, err
    ok, err = _validate_steps('miss')
    if not ok:
        return False, err

    if rule.get('type') == 'dynamic':
        dvs = rule.get('dynamic_values', [])
        if not isinstance(dvs, list) or not dvs:
            return False, 'dynamic_values required for dynamic patterns'
        for x in dvs:
            if not isinstance(x, int) or x < 3 or x > 8:
                return False, 'dynamic_values must be ints between 3 and 8'

    return True, None


def _to_float_or_none(v):
    try:
        if v is None or v == '':
            return None
        return float(v)
    except Exception:
        return None


def _to_int_or_none(v):
    try:
        if v is None or v == '':
            return None
        return int(v)
    except Exception:
        return None


def _flatten_dict_candidates(payload, max_depth=3):
    """Return a list of dict-like candidates (payload + common nested dicts) for key lookup."""
    out = []
    seen = set()

    def _walk(obj, depth):
        if depth < 0:
            return
        if not isinstance(obj, dict):
            return
        oid = id(obj)
        if oid in seen:
            return
        seen.add(oid)
        out.append(obj)

        for k, v in obj.items():
            if isinstance(v, dict):
                _walk(v, depth - 1)
            elif isinstance(v, list) and v and isinstance(v[0], dict):
                # sometimes data is a list of dicts
                _walk(v[0], depth - 1)

    _walk(payload, max_depth)
    return out


def _first_present(candidates, keys):
    for c in candidates:
        for k in keys:
            if isinstance(c, dict) and k in c and c[k] is not None and c[k] != '':
                return c[k]
    return None


def _extract_round_metadata(payload):
    """Best-effort extraction of bc.game style fields from raw payload.
    Keeps ingestion resilient across slightly different API shapes."""
    if not isinstance(payload, dict):
        return {}

    if os.getenv('ROUND_META_DEBUG') == '1':
        try:
            print(f"ROUND_META_DEBUG keys={list(payload.keys())}")
        except Exception:
            pass

    candidates = _flatten_dict_candidates(payload)

    game_id = _first_present(candidates, ['game_id', 'gid', 'gameId', 'game_id_str', 'gameid', 'id'])
    max_rate = _to_float_or_none(_first_present(candidates, ['max_rate', 'maxRate', 'rate_max', 'maxMultiplier', 'max_multiplier', 'rate']))
    total_bet_amount = _to_float_or_none(_first_present(candidates, ['total_bet_amount', 'totalBetAmount', 'total_bet', 'totalBet', 'bet_total']))
    total_win_amount = _to_float_or_none(_first_present(candidates, ['total_win_amount', 'totalWinAmount', 'total_win', 'totalWin', 'win_total']))
    total_profit_amount = _to_float_or_none(_first_present(candidates, ['total_profit_amount', 'totalProfitAmount', 'total_profit', 'totalProfit', 'profit_total']))
    player_count = _to_int_or_none(_first_present(candidates, ['player_count', 'playerCount', 'players', 'player_total']))
    bet_count = _to_int_or_none(_first_present(candidates, ['bet_count', 'betCount', 'bets', 'bet_total_count']))
    hash_value = _first_present(candidates, ['hash', 'hash_value', 'hashValue', 'server_seed_hash', 'serverSeedHash'])
    game_timestamp = _first_present(candidates, ['timestamp', 'ts', 'created_at', 'time', 'ended_at', 'end_time'])
    game_timestamp = str(game_timestamp) if game_timestamp is not None and game_timestamp != '' else None

    return {
        'game_id': str(game_id) if game_id is not None else None,
        'max_rate': max_rate,
        'total_bet_amount': total_bet_amount,
        'total_win_amount': total_win_amount,
        'total_profit_amount': total_profit_amount,
        'player_count': player_count,
        'bet_count': bet_count,
        'hash_value': str(hash_value) if hash_value is not None else None,
        'game_timestamp': game_timestamp,
    }


def _enrich_from_bc_history(value_hint=None):
    """Fetch latest crash object from bc.game history endpoints.
    Returns best-effort metadata dict (same keys as _extract_round_metadata).
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/json',
        'Origin': 'https://bc.game',
        'Referer': 'https://bc.game/game/crash'
    }

    urls = [
        'https://bc.game/api/single/game/crash/history?size=1',
        'https://bc.game/api/game/support/crash/history?size=1',
        'https://bc.game/api/crash/history?limit=1',
    ]

    for url in urls:
        try:
            resp = http_requests.get(url, headers=headers, timeout=10)
            if resp.status_code != 200:
                continue
            data = resp.json()

            if isinstance(data, dict) and 'data' in data:
                data = data['data']

            latest = None
            if isinstance(data, list) and data:
                latest = data[0]
            elif isinstance(data, dict):
                latest = data

            if not isinstance(latest, dict):
                continue

            meta = _extract_round_metadata(latest)
            # sometimes crash value is present; if we have a hint and mismatch is huge, skip
            if value_hint is not None:
                for k in ['value', 'crash', 'multiplier', 'bust', 'point', 'rate', 'v', 'max_rate', 'maxRate']:
                    if k in latest:
                        try:
                            vv = float(latest[k])
                            if abs(vv - float(value_hint)) > 0.01 and vv != float(value_hint):
                                # not necessarily wrong, but try other endpoints before accepting
                                break
                        except Exception:
                            pass

            # derive profit if absent
            if meta.get('total_profit_amount') is None and meta.get('total_win_amount') is not None and meta.get('total_bet_amount') is not None:
                meta['total_profit_amount'] = float(meta['total_win_amount']) - float(meta['total_bet_amount'])

            return meta
        except Exception:
            continue

    return {}

def init_default_patterns():
    with app.app_context():
        if Pattern.query.count() == 0:
            for p in DEFAULT_PATTERNS:
                pattern = Pattern(
                    name=p['name'],
                    description=p.get('description', ''),
                    pattern_type=p.get('type', 'static'),
                    rule_json=json.dumps(p),
                    active=p.get('active', True)
                )
                db.session.add(pattern)
                db.session.flush()
                
                stats = PatternStats(pattern_id=pattern.id)
                db.session.add(stats)
            
            db.session.commit()
            print("Default patterns initialized")

def load_patterns_to_engine():
    with app.app_context():
        patterns = Pattern.query.filter_by(active=True).all()
        pattern_list = []
        
        for p in patterns:
            rule = json.loads(p.rule_json)
            rule['active'] = p.active
            rule['priority'] = p.priority
            
            if rule.get('type') == 'dynamic':
                expanded = pattern_engine.expand_dynamic_pattern(rule)
                for exp in expanded:
                    x = exp.get('dynamic_x')
                    sub_key = f"X={x}" if x is not None else "X=?"
                    pattern_list.append((f"{p.id}:{sub_key}", exp))
            else:
                pattern_list.append((str(p.id), rule))
        
        pattern_engine.load_patterns(pattern_list)
        
        pattern_stats_cache.clear()
        for p in patterns:
            rule = json.loads(p.rule_json) if p.rule_json else {}
            if rule.get('type') == 'dynamic':
                expanded = pattern_engine.expand_dynamic_pattern(rule)
                for exp in expanded:
                    x = exp.get('dynamic_x')
                    sub_key = f"X={x}" if x is not None else "X=?"
                    stat_row = PatternSubStats.query.filter_by(pattern_id=p.id, sub_pattern_key=sub_key).first()
                    if not stat_row:
                        stat_row = PatternSubStats(pattern_id=p.id, sub_pattern_key=sub_key)
                        db.session.add(stat_row)
                        db.session.flush()
                    pattern_stats_cache[f"{p.id}:{sub_key}"] = {
                        'trigger_streak': stat_row.trigger_streak,
                        'miss_streak': stat_row.miss_streak,
                        'total_triggers': stat_row.total_triggers,
                        'total_misses': stat_row.total_misses,
                        'longest_trigger_streak': stat_row.longest_trigger_streak,
                        'longest_miss_streak': stat_row.longest_miss_streak
                    }
            else:
                if p.stats:
                    pattern_stats_cache[str(p.id)] = {
                        'trigger_streak': p.stats.trigger_streak,
                        'miss_streak': p.stats.miss_streak,
                        'total_triggers': p.stats.total_triggers,
                        'total_misses': p.stats.total_misses,
                        'longest_trigger_streak': p.stats.longest_trigger_streak,
                        'longest_miss_streak': p.stats.longest_miss_streak
                    }
        db.session.commit()
        
        print(f"Loaded {len(pattern_list)} patterns")

def load_recent_colors():
    global recent_colors, recent_rounds
    with app.app_context():
        rounds = Round.query.order_by(Round.id.desc()).limit(MAX_COLORS_BUFFER).all()
        recent_colors = [r.color for r in reversed(rounds)]
        recent_rounds = [r.round_number for r in reversed(rounds)]
        print(f"Loaded {len(recent_colors)} recent rounds")

def send_notification(alert_data):
    settings = NotificationSettings.query.first()
    if not settings:
        return
    
    channels = json.loads(settings.enabled_channels) if settings.enabled_channels else ['websocket']
    
    if 'webhook' in channels and settings.webhook_url:
        try:
            http_requests.post(settings.webhook_url, json=alert_data, timeout=5)
        except Exception as e:
            print(f"Webhook notification failed: {e}")
    
    if 'telegram' in channels and settings.telegram_token and settings.telegram_chat_id:
        try:
            msg = f"Pattern Alert: {alert_data['pattern_name']}\nEvent: {alert_data['event']}\nStreak: {alert_data['streak']}"
            url = f"https://api.telegram.org/bot{settings.telegram_token}/sendMessage"
            http_requests.post(url, json={'chat_id': settings.telegram_chat_id, 'text': msg}, timeout=5)
        except Exception as e:
            print(f"Telegram notification failed: {e}")

def check_and_fire_alert(pattern_id, pattern_name, event_type, streak, round_number, rule):
    alerts_config = rule.get('alerts', {})
    trigger_threshold = alerts_config.get('trigger_streak_threshold', DEFAULT_TRIGGER_THRESHOLD)
    miss_threshold = alerts_config.get('miss_threshold', DEFAULT_MISS_THRESHOLD)
    cooldown = alerts_config.get('cooldown_rounds', DEFAULT_COOLDOWN_ROUNDS)
    
    should_alert = False
    if event_type == 'trigger' and streak >= trigger_threshold:
        should_alert = True
        alert_type = 'trigger_streak'
    elif event_type == 'miss' and streak >= miss_threshold:
        should_alert = True
        alert_type = 'miss_streak'
    
    if not should_alert:
        return None
    
    cooldown_key = f"{pattern_id}_{alert_type}"
    if cooldown_key in alert_cooldown_cache:
        if alert_cooldown_cache[cooldown_key] > 0:
            alert_cooldown_cache[cooldown_key] -= 1
            return None
    
    alert_cooldown_cache[cooldown_key] = cooldown
    
    alert_data = {
        'pattern_id': pattern_id,
        'pattern_name': pattern_name,
        'event': alert_type,
        'streak': streak,
        'round': round_number
    }
    
    def _base_pattern_id(pid):
        if isinstance(pid, int):
            return pid
        if not isinstance(pid, str):
            return int(pid)
        if ':' in pid:
            return int(pid.split(':', 1)[0])
        if '_' in pid:
            return int(pid.split('_', 1)[0])
        return int(pid)

    alert = Alert(
        pattern_id=_base_pattern_id(pattern_id),
        alert_type=alert_type,
        payload=json.dumps(alert_data),
        status='sent'
    )
    db.session.add(alert)
    
    socketio.emit('pattern_alert', alert_data, namespace='/')
    
    send_notification(alert_data)
    
    return alert

def save_streak_history(pattern_id, streak_type, streak_length, start_round, end_round, sequence, sub_pattern_key=None):
    if streak_length < 2:
        return

    if isinstance(pattern_id, int):
        base_id = pattern_id
    else:
        pid_str = str(pattern_id)
        if ':' in pid_str:
            base_id = int(pid_str.split(':', 1)[0])
        elif '_' in pid_str:
            base_id = int(pid_str.split('_', 1)[0])
        else:
            base_id = int(pid_str)
    
    streak = StreakHistory(
        pattern_id=base_id,
        streak_type=streak_type,
        streak_length=streak_length,
        sub_pattern_key=sub_pattern_key,
        start_round=start_round,
        end_round=end_round,
        sequence_json=json.dumps(sequence[-streak_length:] if len(sequence) >= streak_length else sequence)
    )
    db.session.add(streak)

def process_round_patterns(round_obj):
    global recent_rounds
    recent_colors.append(round_obj.color)
    recent_rounds.append(round_obj.round_number)
    if len(recent_colors) > MAX_COLORS_BUFFER:
        recent_colors.pop(0)
        recent_rounds.pop(0)
    
    prev_stats = {k: v.copy() for k, v in pattern_stats_cache.items()}
    
    results = pattern_engine.process_round(recent_colors, pattern_stats_cache)
    
    for pattern_key, current_stat in pattern_stats_cache.items():
        prev = prev_stats.get(pattern_key, {})
        prev_trigger = prev.get('trigger_streak', 0)
        prev_miss = prev.get('miss_streak', 0)
        current_trigger = current_stat.get('trigger_streak', 0)
        current_miss = current_stat.get('miss_streak', 0)

        sub_key = None
        if isinstance(pattern_key, str) and ':' in pattern_key:
            sub_key = pattern_key.split(':', 1)[1]
        
        if prev_trigger > 1 and current_trigger < prev_trigger:
            save_streak_history(pattern_key, 'trigger', prev_trigger, 
                               recent_rounds[-prev_trigger-1] if len(recent_rounds) > prev_trigger else '',
                               recent_rounds[-2] if len(recent_rounds) > 1 else '', recent_colors, sub_pattern_key=sub_key)
        
        if prev_miss > 1 and current_miss < prev_miss:
            save_streak_history(pattern_key, 'miss', prev_miss,
                               recent_rounds[-prev_miss-1] if len(recent_rounds) > prev_miss else '',
                               recent_rounds[-2] if len(recent_rounds) > 1 else '', recent_colors, sub_pattern_key=sub_key)
    
    for result in results:
        pattern_key = result['pattern_id']
        pid_str = str(pattern_key)
        base_id = int(pid_str.split(':', 1)[0]) if ':' in pid_str else int(pid_str.split('_', 1)[0]) if '_' in pid_str else int(pid_str)

        stat = result['stats']
        if ':' in pid_str:
            sub_key = pid_str.split(':', 1)[1]
            sub_stat = PatternSubStats.query.filter_by(pattern_id=base_id, sub_pattern_key=sub_key).first()
            if sub_stat:
                sub_stat.trigger_streak = stat['trigger_streak']
                sub_stat.miss_streak = stat['miss_streak']
                sub_stat.total_triggers = stat['total_triggers']
                sub_stat.total_misses = stat['total_misses']
                sub_stat.longest_trigger_streak = stat['longest_trigger_streak']
                sub_stat.longest_miss_streak = stat['longest_miss_streak']
        else:
            pattern_stat = PatternStats.query.filter_by(pattern_id=base_id).first()
            if pattern_stat:
                pattern_stat.trigger_streak = stat['trigger_streak']
                pattern_stat.miss_streak = stat['miss_streak']
                pattern_stat.total_triggers = stat['total_triggers']
                pattern_stat.total_misses = stat['total_misses']
                pattern_stat.longest_trigger_streak = stat['longest_trigger_streak']
                pattern_stat.longest_miss_streak = stat['longest_miss_streak']

        pattern = db.session.get(Pattern, base_id)
        if pattern:
            rule = json.loads(pattern.rule_json)
            check_and_fire_alert(pattern_key, pattern.name, result['event'], 
                                stat['trigger_streak'] if result['event'] == 'trigger' else stat['miss_streak'],
                                round_obj.round_number, rule)
    
    db.session.commit()
    
    round_data = round_obj.to_dict()
    if not live_updates_paused:
        socketio.emit('round_added', round_data, namespace='/')
        
        if results:
            for r in results:
                socketio.emit('pattern_event', r, namespace='/')
    
    return results

def on_crash_received(value, raw_data=None):
    with app.app_context():
        try:
            color = get_color_from_value(value)
            round_number = f"R{int(time.time()*1000)}"

            meta = _extract_round_metadata(raw_data) if raw_data else {}
            if meta.get('max_rate') is None:
                meta['max_rate'] = float(value)

            # If payload didn't include totals/counts/hash, enrich from bc.game REST history.
            if any(meta.get(k) is None for k in ['total_bet_amount', 'total_win_amount', 'player_count', 'bet_count', 'hash_value', 'game_timestamp']):
                enrich = _enrich_from_bc_history(value_hint=value)
                for k, v in enrich.items():
                    if meta.get(k) is None and v is not None:
                        meta[k] = v
            
            round_obj = Round(
                round_number=round_number,
                crash_value=value,
                color=color,
                game_id=meta.get('game_id'),
                max_rate=meta.get('max_rate'),
                total_bet_amount=meta.get('total_bet_amount'),
                total_win_amount=meta.get('total_win_amount'),
                total_profit_amount=meta.get('total_profit_amount'),
                player_count=meta.get('player_count'),
                bet_count=meta.get('bet_count'),
                hash_value=meta.get('hash_value'),
                game_timestamp=meta.get('game_timestamp'),
                raw_payload=json.dumps(raw_data) if raw_data else None
            )
            db.session.add(round_obj)
            db.session.commit()
            
            process_round_patterns(round_obj)
            
            print(f"Round saved: {value}x -> {color}")
            
        except Exception as e:
            print(f"Error processing crash: {e}")
            db.session.rollback()

@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/api/rounds', methods=['GET', 'POST'])
def api_rounds():
    if request.method == 'POST':
        data = request.json
        value = data.get('crash_value', 0)
        color = get_color_from_value(value)
        round_number = data.get('round_number', f"R{int(time.time()*1000)}")

        meta = _extract_round_metadata(data)
        if meta.get('max_rate') is None:
            meta['max_rate'] = float(value)
        
        round_obj = Round(
            round_number=round_number,
            crash_value=value,
            color=color,
            source=data.get('source', 'api'),
            game_id=meta.get('game_id'),
            max_rate=meta.get('max_rate'),
            total_bet_amount=meta.get('total_bet_amount'),
            total_win_amount=meta.get('total_win_amount'),
            total_profit_amount=meta.get('total_profit_amount'),
            player_count=meta.get('player_count'),
            bet_count=meta.get('bet_count'),
            hash_value=meta.get('hash_value'),
            game_timestamp=meta.get('game_timestamp'),
            raw_payload=json.dumps(data)
        )
        db.session.add(round_obj)
        db.session.commit()
        
        process_round_patterns(round_obj)
        
        return jsonify(round_obj.to_dict()), 201
    
    limit = request.args.get('limit', 200, type=int)
    rounds = Round.query.order_by(Round.id.desc()).limit(limit).all()
    return jsonify([r.to_dict() for r in reversed(rounds)])

@app.route('/api/patterns', methods=['GET', 'POST'])
def api_patterns():
    if request.method == 'POST':
        data = request.json
        rule = data.get('rule_json', {})
        if not rule:
            rule = {k: data.get(k) for k in ['name', 'type', 'treat_yellow_as_green', 'trigger', 'miss', 'active']}

        ok, err = _validate_rule_json(rule)
        if not ok:
            return jsonify({'error': err}), 400
        
        pattern = Pattern(
            name=data.get('name', rule.get('name', 'New Pattern')),
            description=data.get('description', ''),
            pattern_type=data.get('pattern_type', rule.get('type', 'static')),
            rule_json=json.dumps(rule),
            active=data.get('active', rule.get('active', True)),
            priority=data.get('priority', 0)
        )
        db.session.add(pattern)
        db.session.flush()
        
        stats = PatternStats(pattern_id=pattern.id)
        db.session.add(stats)
        db.session.commit()
        
        load_patterns_to_engine()
        return jsonify(pattern.to_dict()), 201
    
    patterns = Pattern.query.all()
    result = []
    for p in patterns:
        data = p.to_dict()
        streaks = StreakHistory.query.filter_by(pattern_id=p.id).order_by(StreakHistory.streak_length.desc()).limit(10).all()
        data['longest_streaks'] = {s.streak_length: 1 for s in streaks}
        result.append(data)
    return jsonify(result)

@app.route('/api/patterns/<int:id>', methods=['GET', 'PUT', 'DELETE'])
def api_pattern(id):
    pattern = Pattern.query.get_or_404(id)
    
    if request.method == 'DELETE':
        # Explicit cleanup to avoid FK constraint errors (sqlite/postgres)
        Alert.query.filter_by(pattern_id=pattern.id).delete(synchronize_session=False)
        StreakHistory.query.filter_by(pattern_id=pattern.id).delete(synchronize_session=False)
        PatternSubStats.query.filter_by(pattern_id=pattern.id).delete(synchronize_session=False)
        PatternStats.query.filter_by(pattern_id=pattern.id).delete(synchronize_session=False)
        db.session.delete(pattern)
        db.session.commit()
        load_patterns_to_engine()
        return '', 204
    
    if request.method == 'PUT':
        data = request.json
        pattern.name = data.get('name', pattern.name)
        pattern.description = data.get('description', pattern.description)
        next_rule = data.get('rule_json', json.loads(pattern.rule_json))
        ok, err = _validate_rule_json(next_rule)
        if not ok:
            return jsonify({'error': err}), 400
        pattern.rule_json = json.dumps(next_rule)
        pattern.active = data.get('active', pattern.active)
        if 'priority' in data:
            pattern.priority = data['priority']
        db.session.commit()
        load_patterns_to_engine()
    
    return jsonify(pattern.to_dict())

@app.route('/api/patterns/<int:id>/stats')
def api_pattern_stats(id):
    pattern = Pattern.query.get_or_404(id)
    return jsonify(pattern.stats.to_dict() if pattern.stats else {})

@app.route('/api/analytics/summary')
def api_analytics():
    limit = request.args.get('range', 100, type=int)
    rounds = Round.query.order_by(Round.id.desc()).limit(limit).all()
    
    colors = {'red': 0, 'green': 0, 'yellow': 0}
    for r in rounds:
        colors[r.color] = colors.get(r.color, 0) + 1
    
    total = len(rounds)
    return jsonify({
        'total_rounds': total,
        'distribution': colors,
        'percentages': {k: round(v/total*100, 2) if total > 0 else 0 for k, v in colors.items()}
    })

@app.route('/api/streaks/top')
def api_top_streaks():
    streak_type = request.args.get('type', 'trigger')
    limit = request.args.get('limit', 10, type=int)
    
    streaks = StreakHistory.query.filter_by(streak_type=streak_type)\
        .order_by(StreakHistory.streak_length.desc())\
        .limit(limit).all()
    
    return jsonify([s.to_dict() for s in streaks])

@app.route('/api/patterns/<int:id>/streak-distribution')
def api_pattern_streak_distribution(id):
    pattern = Pattern.query.get_or_404(id)
    streak_type = request.args.get('type', 'trigger')
    
    streaks = StreakHistory.query.filter_by(pattern_id=id, streak_type=streak_type).all()
    
    distribution = {}
    for s in streaks:
        length = s.streak_length
        if length not in distribution:
            distribution[length] = {'total': 0, 'recent': None}
        distribution[length]['total'] += 1
        if distribution[length]['recent'] is None or s.created_at > distribution[length]['recent']:
            distribution[length]['recent'] = s.created_at
    
    result = []
    for length in sorted(distribution.keys(), reverse=True)[:10]:
        result.append({
            'streak': length,
            'total': distribution[length]['total'],
            'recent': distribution[length]['recent'].isoformat() if distribution[length]['recent'] else None
        })
    
    return jsonify({
        'pattern_id': id,
        'pattern_name': pattern.name,
        'streak_type': streak_type,
        'distribution': result
    })

def normalize_dsl_step(step):
    """Normalize a DSL step to ensure consistent schema for frontend rendering"""
    if not isinstance(step, dict):
        return {'label': str(step), 'type': 'unknown'}
    
    colors = step.get('colors', [])
    operator = step.get('operator', 'minimum')
    count = step.get('count', 1)
    
    if 'sequence' in step:
        return {
            'type': 'sequence',
            'label': 'Sequence',
            'steps': [normalize_dsl_step(s) for s in step.get('sequence', [])]
        }
    
    if 'exclude' in step:
        return {
            'type': 'exclusion',
            'label': f"Exclude: {', '.join(step.get('exclude', []))}"
        }
    
    if 'red' in colors:
        color_class = 'red'
        color_label = 'R'
    elif 'yellow' in colors and 'green_yellow' not in colors:
        color_class = 'yellow'
        color_label = 'Y'
    elif 'green' in colors and 'green_yellow' not in colors:
        color_class = 'green'
        color_label = 'G'
    else:
        color_class = 'green'
        color_label = 'G/Y'
    
    op_symbol = '=' if operator == 'exact' else '>=' if operator == 'minimum' else '<='
    count_display = str(count).replace('$', '') if isinstance(count, str) else str(count)
    
    return {
        'type': 'count',
        'color': color_class,
        'operator': operator,
        'count': count,
        'label': f"{op_symbol}{count_display} {color_label}"
    }

def build_visual_sequence(rule):
    """Build visual sequence representation with normalized steps for frontend rendering"""
    trigger_steps = rule.get('trigger', {}).get('steps', [])
    miss_steps = rule.get('miss', {}).get('steps', [])
    
    return {
        'trigger': [normalize_dsl_step(s) for s in trigger_steps],
        'miss': [normalize_dsl_step(s) for s in miss_steps],
        'treat_yellow_as_green': rule.get('treat_yellow_as_green', True),
        'type': rule.get('type', 'static'),
        'dynamic_values': rule.get('dynamic_values', [])
    }

def build_visual_sequence_flat(rule):
    """Return a flat static list of colors for card preview (render-only).
    This intentionally keeps the sequence compact for UI."""
    if rule.get('type') == 'dynamic':
        try:
            expanded = pattern_engine.expand_dynamic_pattern(rule)
            if expanded:
                expanded.sort(key=lambda r: r.get('dynamic_x', 0))
                rule = expanded[0]
        except Exception:
            pass

    steps = rule.get('trigger', {}).get('steps', [])
    out = []
    for step in steps:
        colors = step.get('colors', [])
        op = step.get('operator', 'minimum')
        count = step.get('count', 1)
        try:
            count_int = int(count)
        except Exception:
            count_int = 1

        if 'red' in colors:
            c = 'red'
        elif 'yellow' in colors and 'green_yellow' not in colors:
            c = 'yellow'
        else:
            c = 'green'

        # Render exact counts for clarity (client wants explicit sequence).
        # Safety cap prevents accidental huge arrays.
        repeats = 1
        if op in ['exact', 'minimum']:
            repeats = max(count_int, 1)
        repeats = min(repeats, 25)
        out.extend([c] * repeats)
    return out

@app.route('/api/patterns/dashboard')
def api_patterns_dashboard():
    patterns = Pattern.query.all()
    result = []
    
    from sqlalchemy import func
    trigger_dist_query = db.session.query(
        StreakHistory.pattern_id,
        StreakHistory.streak_length,
        func.count(StreakHistory.id).label('count')
    ).filter(StreakHistory.streak_type == 'trigger')\
     .group_by(StreakHistory.pattern_id, StreakHistory.streak_length).all()

    miss_dist_query = db.session.query(
        StreakHistory.pattern_id,
        StreakHistory.streak_length,
        func.count(StreakHistory.id).label('count')
    ).filter(StreakHistory.streak_type == 'miss')\
     .group_by(StreakHistory.pattern_id, StreakHistory.streak_length).all()

    trigger_dist_map = {}
    for pid, length, cnt in trigger_dist_query:
        trigger_dist_map.setdefault(pid, {})[length] = cnt

    miss_dist_map = {}
    for pid, length, cnt in miss_dist_query:
        miss_dist_map.setdefault(pid, {})[length] = cnt
    
    for p in patterns:
        rule = json.loads(p.rule_json) if p.rule_json else {}

        trigger_dist = trigger_dist_map.get(p.id, {})
        miss_dist = miss_dist_map.get(p.id, {})

        streak_table = []
        # Dynamic: compute top 12 longest unique streak lengths separately for trigger and miss
        top_trigger_lengths = sorted(set(trigger_dist.keys()), reverse=True)[:12]
        top_miss_lengths = sorted(set(miss_dist.keys()), reverse=True)[:12]

        max_rows = max(len(top_trigger_lengths), len(top_miss_lengths), 12)
        max_rows = min(max_rows, 12)

        for i in range(max_rows):
            t_len = top_trigger_lengths[i] if i < len(top_trigger_lengths) else None
            m_len = top_miss_lengths[i] if i < len(top_miss_lengths) else None
            streak_table.append({
                'trigger_streak': int(t_len) if t_len is not None else None,
                'trigger_count': int(trigger_dist.get(t_len, 0)) if t_len is not None else 0,
                'miss_streak': int(m_len) if m_len is not None else None,
                'miss_count': int(miss_dist.get(m_len, 0)) if m_len is not None else 0
            })

        # Flat mapping (trigger distribution) with no hardcoded streak keys
        longest_trigger = {str(int(length)): int(trigger_dist.get(length, 0)) for length in top_trigger_lengths}

        if rule.get('type') == 'dynamic':
            subs = PatternSubStats.query.filter_by(pattern_id=p.id).all()
            current_trigger = max((s.trigger_streak for s in subs), default=0)
            current_miss = max((s.miss_streak for s in subs), default=0)
            max_trigger = max((s.longest_trigger_streak for s in subs), default=0)
            max_miss = max((s.longest_miss_streak for s in subs), default=0)
        else:
            stats = p.stats
            current_trigger = stats.trigger_streak if stats else 0
            current_miss = stats.miss_streak if stats else 0
            max_trigger = stats.longest_trigger_streak if stats else 0
            max_miss = stats.longest_miss_streak if stats else 0

        result.append({
            'id': f"p{p.id}",
            'name': p.name,
            'active': p.active,
            'visual_sequence': build_visual_sequence_flat(rule),
            'stats': {
                'current_trigger': current_trigger,
                'current_miss': current_miss,
                'max_trigger': max_trigger,
                'max_miss': max_miss
            },
            # Client spec expects a flat mapping (trigger distribution)
            'longest_streaks': longest_trigger,
            'streak_table': streak_table
        })
    
    return jsonify({'patterns': result})

@app.route('/api/connection/status')
def api_connection_status():
    return jsonify({
        'connected': bc_listener is not None and bc_listener.running,
        'last_crash': bc_listener.last_crash_value if bc_listener else None
    })

@app.route('/api/alerts', methods=['GET'])
def api_alerts():
    limit = request.args.get('limit', 50, type=int)
    pattern_id = request.args.get('pattern_id', type=int)
    
    query = Alert.query.order_by(Alert.created_at.desc())
    if pattern_id:
        query = query.filter_by(pattern_id=pattern_id)
    
    alerts = query.limit(limit).all()
    return jsonify([a.to_dict() for a in alerts])

@app.route('/api/alerts/clear', methods=['POST'])
def api_clear_alerts():
    Alert.query.delete()
    db.session.commit()
    return jsonify({'status': 'cleared'})

@app.route('/api/colors/config', methods=['GET', 'PUT'])
def api_color_config():
    config = ColorConfig.query.first()
    
    if request.method == 'PUT':
        data = request.json
        if not config:
            config = ColorConfig()
            db.session.add(config)
        
        config.red_max = data.get('red_max', config.red_max)
        config.green_max = data.get('green_max', config.green_max)
        db.session.commit()
        invalidate_color_config_cache()
        return jsonify(config.to_dict())
    
    if not config:
        return jsonify({
            'red_max': 2.0,
            'green_max': 10.0,
            'rules': {
                'red': 'value < 2.0',
                'green': '2.0 <= value < 10.0',
                'yellow': 'value >= 10.0'
            }
        })
    
    return jsonify(config.to_dict())

@app.route('/api/notifications/settings', methods=['GET', 'PUT'])
def api_notification_settings():
    settings = NotificationSettings.query.first()
    
    if request.method == 'PUT':
        data = request.json
        if not settings:
            settings = NotificationSettings()
            db.session.add(settings)
        
        if 'telegram_token' in data:
            settings.telegram_token = data['telegram_token']
        if 'telegram_chat_id' in data:
            settings.telegram_chat_id = data['telegram_chat_id']
        if 'webhook_url' in data:
            settings.webhook_url = data['webhook_url']
        if 'email_smtp_host' in data:
            settings.email_smtp_host = data['email_smtp_host']
        if 'email_smtp_port' in data:
            settings.email_smtp_port = data['email_smtp_port']
        if 'email_username' in data:
            settings.email_username = data['email_username']
        if 'email_password' in data:
            settings.email_password = data['email_password']
        if 'email_to' in data:
            settings.email_to = data['email_to']
        if 'enabled_channels' in data:
            settings.enabled_channels = json.dumps(data['enabled_channels'])
        
        db.session.commit()
        return jsonify(settings.to_dict())
    
    if not settings:
        return jsonify({
            'telegram_enabled': False,
            'webhook_enabled': False,
            'email_enabled': False,
            'enabled_channels': ['websocket']
        })
    
    return jsonify(settings.to_dict())

@app.route('/api/patterns/<int:id>/simulate', methods=['POST'])
def api_pattern_simulate(id):
    pattern = Pattern.query.get_or_404(id)
    rule = json.loads(pattern.rule_json)
    
    data = request.json or {}
    limit = data.get('limit', 500)
    
    rounds = Round.query.order_by(Round.id.asc()).limit(limit).all()
    colors = [r.color for r in rounds]
    
    if rule.get('type') == 'dynamic':
        expanded = pattern_engine.expand_dynamic_pattern(rule)
        rules_to_test = [(f"{id}_{i}", exp) for i, exp in enumerate(expanded)]
    else:
        rules_to_test = [(id, rule)]
    
    results = {
        'pattern_name': pattern.name,
        'total_rounds': len(rounds),
        'sub_patterns': []
    }
    
    for pid, r in rules_to_test:
        triggers = 0
        misses = 0
        trigger_streak = 0
        miss_streak = 0
        max_trigger_streak = 0
        max_miss_streak = 0
        
        for i in range(1, len(colors) + 1):
            window = colors[:i]
            check = pattern_engine.check_pattern(window, r)
            
            if check['trigger']:
                triggers += 1
                trigger_streak += 1
                miss_streak = 0
                if trigger_streak > max_trigger_streak:
                    max_trigger_streak = trigger_streak
            elif check['miss']:
                misses += 1
                miss_streak += 1
                trigger_streak = 0
                if miss_streak > max_miss_streak:
                    max_miss_streak = miss_streak
        
        results['sub_patterns'].append({
            'name': r.get('name', pattern.name),
            'total_triggers': triggers,
            'total_misses': misses,
            'max_trigger_streak': max_trigger_streak,
            'max_miss_streak': max_miss_streak,
            'trigger_rate': round(triggers / len(rounds) * 100, 2) if rounds else 0,
            'miss_rate': round(misses / len(rounds) * 100, 2) if rounds else 0
        })
    
    return jsonify(results)

live_updates_paused = False

@app.route('/api/live/pause', methods=['POST'])
def api_pause_live():
    global live_updates_paused
    live_updates_paused = True
    return jsonify({'status': 'paused', 'paused': True})

@app.route('/api/live/resume', methods=['POST'])
def api_resume_live():
    global live_updates_paused
    live_updates_paused = False
    return jsonify({'status': 'resumed', 'paused': False})

@app.route('/api/live/status')
def api_live_status():
    return jsonify({'paused': live_updates_paused})

@app.route('/api/rounds/history')
def api_rounds_history():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    color = request.args.get('color')
    min_value = request.args.get('min_value', type=float)
    max_value = request.args.get('max_value', type=float)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = Round.query
    
    if color:
        query = query.filter(Round.color == color)
    if min_value is not None:
        query = query.filter(Round.crash_value >= min_value)
    if max_value is not None:
        query = query.filter(Round.crash_value <= max_value)
    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            query = query.filter(Round.created_at >= dt_from)
        except:
            pass
    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            query = query.filter(Round.created_at <= dt_to)
        except:
            pass
    
    total = query.count()
    rounds = query.order_by(Round.id.desc()).offset((page - 1) * per_page).limit(per_page).all()
    
    return jsonify({
        'rounds': [r.to_dict() for r in rounds],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })

@app.route('/api/export/data')
def api_export_data():
    limit = request.args.get('limit', 1000, type=int)
    color = request.args.get('color')
    
    query = Round.query
    if color:
        query = query.filter(Round.color == color)
    
    rounds = query.order_by(Round.id.desc()).limit(limit).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Crash Data"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    
    headers = ['ID', 'Round Number', 'Crash Value', 'Color', 'Source', 'Game ID', 'Max Rate', 'Total Bet', 'Total Win', 'Total Profit', 'Players', 'Bets', 'Hash', 'Timestamp', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    color_fills = {
        'red': PatternFill(start_color="ff4757", end_color="ff4757", fill_type="solid"),
        'green': PatternFill(start_color="2ed573", end_color="2ed573", fill_type="solid"),
        'yellow': PatternFill(start_color="feca57", end_color="feca57", fill_type="solid")
    }
    
    for row, r in enumerate(rounds, 2):
        ws.cell(row=row, column=1, value=r.id)
        ws.cell(row=row, column=2, value=r.round_number)
        ws.cell(row=row, column=3, value=r.crash_value)
        color_cell = ws.cell(row=row, column=4, value=r.color.upper())
        if r.color in color_fills:
            color_cell.fill = color_fills[r.color]
            color_cell.font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=5, value=r.source)
        ws.cell(row=row, column=6, value=r.game_id)
        ws.cell(row=row, column=7, value=r.max_rate)
        ws.cell(row=row, column=8, value=r.total_bet_amount)
        ws.cell(row=row, column=9, value=r.total_win_amount)
        ws.cell(row=row, column=10, value=r.total_profit_amount)
        ws.cell(row=row, column=11, value=r.player_count)
        ws.cell(row=row, column=12, value=r.bet_count)
        ws.cell(row=row, column=13, value=r.hash_value)
        ws.cell(row=row, column=14, value=r.game_timestamp)
        ws.cell(row=row, column=15, value=r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '')
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'crash_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/api/export/patterns')
def api_export_patterns():
    patterns = Pattern.query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pattern Analysis"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    
    headers = ['Pattern Name', 'Type', 'Active', 'Current Trigger', 'Current Miss', 
               'Total Triggers', 'Total Misses', 'Max Trigger Streak', 'Max Miss Streak',
               'Trigger Rate %', 'Miss Rate %']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    total_rounds = Round.query.count()
    
    for row, p in enumerate(patterns, 2):
        stats = p.stats
        ws.cell(row=row, column=1, value=p.name)
        ws.cell(row=row, column=2, value=p.pattern_type)
        ws.cell(row=row, column=3, value='Yes' if p.active else 'No')
        ws.cell(row=row, column=4, value=stats.trigger_streak if stats else 0)
        ws.cell(row=row, column=5, value=stats.miss_streak if stats else 0)
        ws.cell(row=row, column=6, value=stats.total_triggers if stats else 0)
        ws.cell(row=row, column=7, value=stats.total_misses if stats else 0)
        ws.cell(row=row, column=8, value=stats.longest_trigger_streak if stats else 0)
        ws.cell(row=row, column=9, value=stats.longest_miss_streak if stats else 0)
        
        if stats and total_rounds > 0:
            ws.cell(row=row, column=10, value=round((stats.total_triggers / total_rounds) * 100, 2))
            ws.cell(row=row, column=11, value=round((stats.total_misses / total_rounds) * 100, 2))
        else:
            ws.cell(row=row, column=10, value=0)
            ws.cell(row=row, column=11, value=0)
    
    ws2 = wb.create_sheet(title="Streak History")
    streak_headers = ['Pattern Name', 'Streak Type', 'Streak Length', 'Start Round', 'End Round', 'Date']
    for col, header in enumerate(streak_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    streaks = StreakHistory.query.order_by(StreakHistory.streak_length.desc()).limit(100).all()
    for row, s in enumerate(streaks, 2):
        ws2.cell(row=row, column=1, value=s.pattern.name if s.pattern else 'Unknown')
        ws2.cell(row=row, column=2, value=s.streak_type)
        ws2.cell(row=row, column=3, value=s.streak_length)
        ws2.cell(row=row, column=4, value=s.start_round)
        ws2.cell(row=row, column=5, value=s.end_round)
        ws2.cell(row=row, column=6, value=s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else '')
    
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'pattern_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/api/export/chart')
def api_export_chart():
    chart_type = request.args.get('type', 'distribution')
    limit = request.args.get('limit', 200, type=int)
    
    rounds = Round.query.order_by(Round.id.desc()).limit(limit).all()
    
    if chart_type == 'distribution':
        colors = {'red': 0, 'green': 0, 'yellow': 0}
        for r in rounds:
            colors[r.color] = colors.get(r.color, 0) + 1
        
        fig, ax = plt.subplots(figsize=(10, 8))
        fig.patch.set_facecolor('#0f0f1a')
        ax.set_facecolor('#0f0f1a')
        
        labels = ['Red (<2x)', 'Green (2-10x)', 'Yellow (>=10x)']
        sizes = [colors['red'], colors['green'], colors['yellow']]
        chart_colors = ['#ff4757', '#2ed573', '#feca57']
        explode = (0.05, 0.05, 0.05)
        
        wedges, texts, autotexts = ax.pie(sizes, explode=explode, labels=labels, colors=chart_colors,
                                          autopct='%1.1f%%', shadow=True, startangle=90)
        
        for text in texts:
            text.set_color('white')
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
        
        ax.set_title(f'Color Distribution (Last {limit} Rounds)', color='white', fontsize=14, fontweight='bold')
        
    elif chart_type == 'timeline':
        fig, ax = plt.subplots(figsize=(14, 6))
        fig.patch.set_facecolor('#0f0f1a')
        ax.set_facecolor('#1a1a2e')
        
        values = [r.crash_value for r in reversed(rounds)]
        colors_list = ['#ff4757' if v < 2 else '#2ed573' if v < 10 else '#feca57' for v in values]
        
        ax.bar(range(len(values)), values, color=colors_list, width=0.8)
        ax.set_xlabel('Round', color='white')
        ax.set_ylabel('Crash Value', color='white')
        ax.set_title(f'Crash Values Timeline (Last {limit} Rounds)', color='white', fontsize=14, fontweight='bold')
        ax.tick_params(colors='white')
        ax.spines['bottom'].set_color('white')
        ax.spines['left'].set_color('white')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        
    elif chart_type == 'patterns':
        patterns = Pattern.query.all()
        
        fig, ax = plt.subplots(figsize=(12, 8))
        fig.patch.set_facecolor('#0f0f1a')
        ax.set_facecolor('#1a1a2e')
        
        names = [p.name for p in patterns]
        triggers = [p.stats.total_triggers if p.stats else 0 for p in patterns]
        misses = [p.stats.total_misses if p.stats else 0 for p in patterns]
        
        x = range(len(names))
        width = 0.35
        
        bars1 = ax.bar([i - width/2 for i in x], triggers, width, label='Triggers', color='#2ed573')
        bars2 = ax.bar([i + width/2 for i in x], misses, width, label='Misses', color='#ff4757')
        
        ax.set_xlabel('Pattern', color='white')
        ax.set_ylabel('Count', color='white')
        ax.set_title('Pattern Triggers vs Misses', color='white', fontsize=14, fontweight='bold')
        ax.set_xticks(list(x))
        ax.set_xticklabels(names, rotation=45, ha='right', color='white')
        ax.tick_params(colors='white')
        ax.legend(facecolor='#1a1a2e', edgecolor='white', labelcolor='white')
        ax.spines['bottom'].set_color('white')
        ax.spines['left'].set_color('white')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        
        plt.tight_layout()
    
    else:
        return jsonify({'error': 'Invalid chart type'}), 400
    
    output = io.BytesIO()
    plt.savefig(output, format='png', dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='image/png',
        as_attachment=True,
        download_name=f'chart_{chart_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'
    )

@app.route('/api/analytics/detailed')
def api_detailed_analytics():
    limit = request.args.get('range', 500, type=int)
    rounds = Round.query.order_by(Round.id.desc()).limit(limit).all()
    
    colors = {'red': 0, 'green': 0, 'yellow': 0}
    values = []
    hourly_dist = {}
    streaks = {'red': [], 'green': [], 'yellow': []}
    current_streak = {'color': None, 'count': 0}
    
    for r in rounds:
        colors[r.color] = colors.get(r.color, 0) + 1
        values.append(r.crash_value)
        
        if r.created_at:
            hour = r.created_at.hour
            if hour not in hourly_dist:
                hourly_dist[hour] = {'red': 0, 'green': 0, 'yellow': 0}
            hourly_dist[hour][r.color] += 1
        
        if r.color == current_streak['color']:
            current_streak['count'] += 1
        else:
            if current_streak['color'] and current_streak['count'] > 1:
                streaks[current_streak['color']].append(current_streak['count'])
            current_streak = {'color': r.color, 'count': 1}
    
    if current_streak['color'] and current_streak['count'] > 1:
        streaks[current_streak['color']].append(current_streak['count'])
    
    total = len(rounds)
    avg_value = sum(values) / len(values) if values else 0
    max_value = max(values) if values else 0
    min_value = min(values) if values else 0
    
    def calc_avg_streak(s_list):
        return sum(s_list) / len(s_list) if s_list else 0
    
    return jsonify({
        'total_rounds': total,
        'distribution': colors,
        'percentages': {k: round(v/total*100, 2) if total > 0 else 0 for k, v in colors.items()},
        'value_stats': {
            'average': round(avg_value, 2),
            'max': max_value,
            'min': min_value
        },
        'streak_stats': {
            'red': {
                'max': max(streaks['red']) if streaks['red'] else 0,
                'avg': round(calc_avg_streak(streaks['red']), 2),
                'count': len(streaks['red'])
            },
            'green': {
                'max': max(streaks['green']) if streaks['green'] else 0,
                'avg': round(calc_avg_streak(streaks['green']), 2),
                'count': len(streaks['green'])
            },
            'yellow': {
                'max': max(streaks['yellow']) if streaks['yellow'] else 0,
                'avg': round(calc_avg_streak(streaks['yellow']), 2),
                'count': len(streaks['yellow'])
            }
        },
        'hourly_distribution': hourly_dist
    })

@socketio.on('connect')
def handle_connect():
    print("Client connected")
    emit('connection_status', {'status': 'connected'})

@socketio.on('disconnect')
def handle_disconnect():
    print("Client disconnected")

def start_bc_listener():
    global bc_listener
    bc_listener = BCGameListener(on_crash_callback=on_crash_received)
    bc_listener.start()

with app.app_context():
    db.create_all()

    # lightweight schema upgrades for sqlite (dev)
    try:
        if db.engine.dialect.name == 'sqlite':
            # rounds metadata columns
            round_cols = [row[1] for row in db.session.execute(text("PRAGMA table_info(rounds)")).fetchall()]
            round_add_cols = [
                ('game_id', 'VARCHAR(100)'),
                ('max_rate', 'FLOAT'),
                ('total_bet_amount', 'FLOAT'),
                ('total_win_amount', 'FLOAT'),
                ('total_profit_amount', 'FLOAT'),
                ('player_count', 'INTEGER'),
                ('bet_count', 'INTEGER'),
                ('hash_value', 'VARCHAR(200)'),
                ('game_timestamp', 'VARCHAR(100)')
            ]
            for col_name, col_type in round_add_cols:
                if col_name not in round_cols:
                    db.session.execute(text(f"ALTER TABLE rounds ADD COLUMN {col_name} {col_type}"))
            db.session.commit()

            cols = [row[1] for row in db.session.execute(text("PRAGMA table_info(streak_history)")).fetchall()]
            if 'sub_pattern_key' not in cols:
                db.session.execute(text("ALTER TABLE streak_history ADD COLUMN sub_pattern_key VARCHAR(50)"))
                db.session.commit()

            db.session.execute(text("""
                CREATE TABLE IF NOT EXISTS pattern_sub_stats (
                    id INTEGER PRIMARY KEY,
                    pattern_id INTEGER NOT NULL,
                    sub_pattern_key VARCHAR(50) NOT NULL,
                    trigger_streak INTEGER DEFAULT 0,
                    miss_streak INTEGER DEFAULT 0,
                    total_triggers INTEGER DEFAULT 0,
                    total_misses INTEGER DEFAULT 0,
                    longest_trigger_streak INTEGER DEFAULT 0,
                    longest_miss_streak INTEGER DEFAULT 0,
                    updated_at DATETIME,
                    CONSTRAINT uq_pattern_sub_stats_pattern_key UNIQUE (pattern_id, sub_pattern_key),
                    FOREIGN KEY(pattern_id) REFERENCES patterns(id)
                )
            """))
            db.session.commit()
    except Exception as e:
        print(f"Schema upgrade skipped/failed: {e}")

    init_default_patterns()
    load_patterns_to_engine()
    load_recent_colors()

listener_thread = threading.Thread(target=start_bc_listener, daemon=True)
listener_thread.start()

if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000, debug=False)
